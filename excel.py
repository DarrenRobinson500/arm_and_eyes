import openpyxl as xl
from scene import *


def add_model_to_excel(scene):
    file = "excel/xyz.xlsx"
    wb = xl.load_workbook(file)
    sheet = wb['main']
    start_row = 5
    row = start_row
    column = 1
    var_strings = ["x1", "y1", "x2", "y2", "x1 / (x2 + d)", "x1 / (y2 + d)", "y1 / (x2 + d)", "y1 / (y2 + d)", "x2 / (x1 + d)", "x2 / (y1 + d)", "y2 / (x1 + d)", "y2 / (y1 + d)"]
    sheet.cell(start_row - 1, column).value = "c"
    for var in var_strings:
        # sheet.cell(row, column).value = str("'" + var)
        sheet.cell(row, column).value = var
        row += 1
    column += 1
    row = start_row

    for model in [scene.model_x, scene.model_y, scene.model_z]:
        sheet.cell(start_row - 1, column).value = model.intercept_
        for coef in model.coef_:
            sheet.cell(row, column).value = coef
            row += 1
        column += 1
        row = start_row
    wb.save(file)
    wb.close()

def add_points_to_excel(points):
    file = "excel/xyz.xlsx"
    wb = xl.load_workbook(file)
    sheet = wb['main']
    start_row = 5
    row = start_row
    column = 7
    for point in points:
        sheet.cell(row - 2, column).value = 100
        sheet.cell(row - 1, column).value = 1
        for data in point.data:
            sheet.cell(row, column).value = data
            row += 1
            if row == start_row + 4: row += 13
        column += 1
        row = start_row
    wb.save(file)
    wb.close()


def print_model_x(scene):
    var_strings = ["x1", "y1", "x2", "y2", "x1 / (x2 + d)", "x1 / (y2 + d)", "y1 / (x2 + d)", "y1 / (y2 + d)", "x2 / (x1 + d)", "x2 / (y1 + d)", "y2 / (x1 + d)", "y2 / (y1 + d)"]
    text = ""
    for coef, var_string in zip(scene.model_x.coef_, var_strings):
        text += str(round(coef,1)) + var_string + " + "
    print("Model X:", text[:-3])


# add_model_to_excel(scenes[0])
# add_points_to_excel(scenes[0].calibration_points)